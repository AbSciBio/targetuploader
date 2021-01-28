from openpyxl import load_workbook, Workbook
import parser
from ptdb_request import PTDBRequest

loc = './assets/XOLO_PTDB_Nomenclature.xlsx'
wb = load_workbook(loc)
ws = wb.active
max_row = ws.max_row

results_workbook = Workbook()
results_workbook.active
results_worksheet = results_workbook.create_sheet('results')
results_worksheet.title = "Results of CD19 Upload"
results_worksheet_column_names = [ 
    "Target", 
    "Notes", 
    "Partner", 
    "Project", 
    "Subunit 1 Name", 
    "Subunit 1 Fasta File IDs", 
    "Subunit 1 AA sequence", 
    "Subunit 1 DNA Sequence",
    "Subunit 2 Name", 
    "Subunit 2 Fasta File IDs", 
    "Subunit 2 AA sequence", 
    "Subunit 2 DNA Sequence" 
]
for i in range(0, len(results_worksheet_column_names)):
    results_worksheet.cell(1, i + 1, results_worksheet_column_names[i])

targets = []

for row in range(2, max_row + 1):
    target = {
        "target": ws.cell(row, 1).value,
        "partner": "",
        "protein_class_pk": "1",
        "notes": ws.cell(row, 2).value,
        "project_name": "Xolo",
        "subunits": parser.SubunitParser(ws, row).subunits,
    }
    PTDBRequest(target).post_target()
    targets.append(target)

for i in range(0, len(targets)):
    print(targets[i]['target'])
    results_worksheet.cell(i + 2, 1, targets[i]['target'])
    results_worksheet.cell(i + 2, 2, targets[i]['notes'])
    results_worksheet.cell(i + 2, 3, targets[i]['partner'])
    results_worksheet.cell(i + 2, 4, targets[i]['project_name'])
    results_worksheet.cell(i + 2, 5, targets[i]['subunits'][0]['subunit_name'])
    results_worksheet.cell(i + 2, 6, targets[i]['subunits'][0]['subunit_name'])
    results_worksheet.cell(i + 2, 7, targets[i]['subunits'][0]['amino_acid_sequence'])
    results_worksheet.cell(i + 2, 8, targets[i]['subunits'][0]['genes'][0]['dna_sequence'])
    results_worksheet.cell(i + 2, 9, targets[i]['subunits'][1]['subunit_name'])
    results_worksheet.cell(i + 2, 10, targets[i]['subunits'][1]['subunit_name'])
    results_worksheet.cell(i + 2, 11, targets[i]['subunits'][1]['amino_acid_sequence'])
    results_worksheet.cell(i + 2, 12, targets[i]['subunits'][1]['genes'][0]['dna_sequence'])

results_workbook.save('test.xlsx')