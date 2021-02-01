from Bio import SeqIO
from lib.subunit_parser import SubunitParser
from openpyxl import load_workbook, Workbook

class FastaReader():
  def __init__(self, subunit_name):
    self.file_name = subunit_name 

  def get_fasta_sequence(self, seq_type):
    errors = []
    sequence = ''
    try:
      for record in SeqIO.parse(f'./assets/CD19_MicAbody_Xolo_FASTA_Files/{self.file_name}_{seq_type}.fasta', 'fasta'):
        sequence = str(record.seq)
        
    except FileNotFoundError as err:
      errors.append(f'{err}: {err.filename2}')
   
    if len(sequence) > 0:
      return sequence
    if (errors):
      return errors

  
  
class ExcelReader():
  def __init__(self, excel_path):
    wb = load_workbook(excel_path)
    self.ws = wb.active
    self.max_row = self.ws.max_row
    self.targets = []  
  
    for row in range(2, self.max_row + 1):
        target = {
            "target": self.ws.cell(row, 1).value,
            "partner": "",
            "protein_class_pk": "1",
            "notes": self.ws.cell(row, 2).value,
            "project_name": "Xolo",
            "subunits": SubunitParser(self.ws, row).subunits,
        }
        self.targets.append(target)