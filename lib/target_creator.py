from openpyxl import load_workbook, Workbook
from Bio import SeqIO

class TargetCreator():
  def __init__(self, user_input):
    wb = load_workbook(user_input['excel_file_location'])
    self.ws = wb.active
    self.max_col = self.ws.max_column
    self.max_row = self.ws.max_row
    self.targets = []
    self.fasta_location = user_input['fasta_location']
    self.excl_reader()

  def excl_reader(self):
      for row in range(2, self.max_row + 1):
          target = {
              "target": self.ws.cell(row, 1).value,
              "partner": "Xyphos",
              "protein_class_pk": "1",
              "notes": self.ws.cell(row, 2).value,
              "project_name": "Xolo",
              "subunits": self.parse_subunits(row),
          }
          self.targets.append(target)
    
  def parse_subunits(self, row):
      subunit_array = []
      for i in range(3, self.max_col + 1, 2):
          subunit_name = self.ws.cell(row, i).value 
          subunit = {
              "subunit_name": subunit_name,
              "copies": "1",
              "amino_acid_fasta_description": subunit_name,
              "amino_acid_sequence": self.get_fasta_sequence(subunit_name, "AA"),
              "genes": self.parse_genes(subunit_name)
          }
          subunit_array.append(subunit)
      return subunit_array

  def parse_genes(self, subunit_name):
      gene = {}
      gene["dna_fasta_description"] = subunit_name
      gene["dna_sequence"] = self.get_fasta_sequence(subunit_name, "DNA")
      return [(gene)]

  def get_fasta_sequence(self, file_name, seq_type):
      errors = []
      sequence = ''
      try:
          for record in SeqIO.parse(f'{self.fasta_location}/{file_name}_{seq_type}.fasta', 'fasta'):
              sequence = str(record.seq)      
      except FileNotFoundError as err:
          errors.append(f'{err}: {err.filename2}')
      if len(sequence) > 0:
          return sequence
      if (errors):
          return errors